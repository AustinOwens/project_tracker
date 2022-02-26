#####################################################
## Desc: Makes a sticky note on Desktop using Tkinter
#####################################################
## Dependencies: openpyxl
#####################################################
## Author: Austin Owens
#####################################################

from tkinter import *
import tkinter.scrolledtext as tkst
from tkinter import messagebox
from tkinter import font
import openpyxl

PROJECT_TRACKER_NAME = "Project_Tracker.xlsx"
no_of_windows = 1

class StickyNotes(Toplevel):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.xclick = 0
        self.yclick = 0

        # master (root) window
        self.overrideredirect(True)
        global no_of_windows
        self.geometry('250x250+' + str(1000+no_of_windows*(-30)) + '+' + str(100 + no_of_windows*20))
        self.config(bg = '#838383')
        self.attributes('-topmost', 'true')
        self.resizable(True,True)

        # titlebar
        self.titlebar = Frame(self, bg = '#F8F796', relief = 'flat', bd = 2)
        self.titlebar.bind('<Button-1>', self.get_pos)
        self.titlebar.bind('<B1-Motion>', self.move_window)
        self.titlebar.pack(fill = X, expand = 1, side = TOP)

        self.closebutton = Label(self.titlebar, text = 'X', bg = '#F8F7B6', relief = 'flat')
        self.closebutton.bind('<Button-1>', self.quit_window)
        self.closebutton.pack(side = RIGHT)

        self.newbutton = Label(self.titlebar, text = '+', bg = '#F8F7B6', relief = 'flat')
        self.newbutton.pack(side = LEFT)
        self.newbutton.bind('<Button-1>', self.another_window)

        # main text area
        self.mainarea = tkst.ScrolledText(self, bg = '#FDFDCA', font=('Comic Sans MS', 14, 'italic'), relief = 'flat', padx = 5, pady = 10)
        self.mainarea.pack(fill = BOTH, expand = 1)

        # write in mainarea
        action_items_list = self.get_action_items()
        for index, action_item in enumerate(action_items_list):
            self.mainarea.insert("{}.0".format(index+1), "{}. {}\n".format(index+1, action_item))
            

        # frames to introduce shadows
        self.shadow = Frame(self).pack(side=BOTTOM)
        self.shadow = Frame(self).pack(side=RIGHT)
        
        no_of_windows += 1

    def get_pos(self, event):
        self.xclick = event.x
        self.yclick = event.y

    def move_window(self, event):
        self.geometry('+{0}+{1}'.format(event.x_root-self.xclick, event.y_root-self.yclick))

    def another_window(self, event):
        sticky = StickyNotes(root)

    def quit_window(self, event):
        self.closebutton.config(relief = 'flat', bd = 0)
        if(messagebox.askyesno('Delete Note?','Are you sure you want to delete this note?', parent = self)):
            global no_of_windows
            self.destroy()
            no_of_windows -= 1
            if(no_of_windows == 1):
                root.destroy()
            return
        self.closebutton.config(relief = 'flat', bd = 0, bg = '#F8F7B6')

    def get_action_items(self):
        wb = openpyxl.load_workbook(PROJECT_TRACKER_NAME, data_only=True)
        sheets = wb.sheetnames
        ws = wb[sheets[1]]
        action_item_array = []
        start_reading_flag = False
        
        for row in ws.iter_rows(values_only=True):
            if row[3]:
                if "Action Item" in row[3]:
                    start_reading_flag=True
                
                elif start_reading_flag:
                    if row[6] != 'c':
                        action_item_array.append(row[3])
                        
        return action_item_array
        
if __name__ == "__main__":
    root = Tk()
    root.withdraw()
    # make the first note.
    sticky = StickyNotes(root) 
    root.mainloop()


    
